import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

def extract_values_from_file(file_path):
    first_array = []
    second_array = []
    ra_found = False

    with open(file_path, 'r') as file:
        for line in file:
            line = line.strip()
            if line:
                if "RA" in line:
                    ra_found = True
                    continue
                
                value = line.split()[1] if len(line.split()) > 1 else None
                if value:
                    if not ra_found:
                        first_array.append(value)
                    else:
                        second_array.append(value)

    return first_array, second_array


file_path = 'values.txt'
first_array, second_array = extract_values_from_file(file_path)


input_file = 'sorted_2.xlsx'  
sheet_name = 'Sheet1'  
output_file = 'sorted_3.xlsx'  


if os.path.exists(output_file):
    os.remove(output_file)  


df = pd.read_excel(input_file, sheet_name=sheet_name, header=4)
df['Account Combination'] = df['Account Combination'].astype(str)


sum_row = df.tail(1)
df = df.iloc[:-1]  


wb = load_workbook(input_file)
ws = wb[sheet_name]


row_styles = {}
for row in ws.iter_rows(min_row=6, max_row=ws.max_row, max_col=ws.max_column):
    row_styles[row[0].row] = [(cell.value, cell._style) for cell in row]


ws.delete_rows(6, ws.max_row - 5)



for i, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=6):
    for j, value in enumerate(row, start=1):
        cell = ws.cell(row=i, column=j, value=value)
        if i in row_styles:
            cell._style = row_styles[i][j-1][1]



first_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  
second_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  
default_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")


num_columns = df.shape[1]


for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=1, max_col=num_columns):
    account_combination = row[0].value
    if account_combination:
       
        account_value = account_combination.split('.')[2][:-5]
        
        if account_value in first_array:
            for cell in row:
                cell.fill = first_fill
        elif account_value in second_array:
            for cell in row:
                cell.fill = second_fill
        else:
            for cell in row:
                cell.fill = default_fill

if not sum_row.empty:  
    
    last_row_index = ws.max_row + 1
    for j, value in enumerate(sum_row.values.flatten(), start=1):
        cell = ws.cell(row=last_row_index, column=j, value=value)

       
        fifth_row_style = ws.cell(row=5, column=j)._style
        cell._style = fifth_row_style 


wb.save(output_file)

print(f"Highlighting completed. The highlighted data is saved to {output_file}.")
