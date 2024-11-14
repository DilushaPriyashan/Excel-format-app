import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


input_file = 'input.xlsx' 
sheet_name = 'Sheet1' 
output_file = 'sorted_1.xlsx'  


if os.path.exists(output_file):
    os.remove(output_file)  


df = pd.read_excel(input_file, sheet_name=sheet_name, header=4)


df['Account Combination'] = df['Account Combination'].astype(str)


sum_row = df.tail(1) 


df = df.iloc[:-1]


df['SortKey'] = df['Account Combination'].apply(lambda x: x.split('.')[3] if len(x.split('.')) > 3 else '')

df_sorted = df.sort_values(by='SortKey')

df_sorted = df_sorted.drop(columns=['SortKey'])

df_sorted = pd.concat([df_sorted, sum_row])

wb = load_workbook(input_file)
ws = wb[sheet_name]


row_styles = {}
for row in ws.iter_rows(min_row=6, max_row=ws.max_row, max_col=ws.max_column):
    row_styles[row[0].row] = [(cell.value, cell._style) for cell in row]


ws.delete_rows(6, ws.max_row - 5)


for i, row in enumerate(dataframe_to_rows(df_sorted, index=False, header=False), start=6):
    for j, value in enumerate(row, start=1):
        cell = ws.cell(row=i, column=j, value=value)
        if i in row_styles:
            
            cell._style = row_styles[i][j-1][1]


wb.save(output_file)

print(f"Sorting completed. The sorted data (excluding the sum row) is saved to {output_file}.")
