import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill


def extract_values_from_file(file_path):
    first_array = []
    second_array = []
    ra_found = False

    with open(file_path, 'r') as file:
        for line in file:
            line = line.strip()  
            if line:  
                if "RA" in line:
                    ra_found = True
                    continue  
                
                value = line.split()[1] if len(line.split()) > 1 else None
                if value:
                    if not ra_found:
                        first_array.append(value)
                    else:
                        second_array.append(value)

    return first_array, second_array


file_path = 'values.txt'
first_array, second_array = extract_values_from_file(file_path)


input_file = 'sorted_1.xlsx'  
sheet_name = 'Sheet1'  
output_file = 'sorted_2.xlsx'  


if os.path.exists(output_file):
    os.remove(output_file)  


df = pd.read_excel(input_file, sheet_name=sheet_name, header=4)


df['Account Combination'] = df['Account Combination'].astype(str)

sum_row = df.tail(1)
df = df.iloc[:-1]  


def get_third_value_part(account_comb):
    third_value = account_comb.split('.')[2]  
    return third_value[:-5]  


def shuffle_group(group):
    first_set = []
    second_set = []
    remaining_set = []

    for index, row in group.iterrows():
        third_value_part = get_third_value_part(row['Account Combination'])
        if third_value_part in first_array:
            first_set.append(row)
        elif third_value_part in second_array:
            second_set.append(row)
        else:
            remaining_set.append(row)

    
    return pd.DataFrame(first_set + second_set + remaining_set)


df['4th_value'] = df['Account Combination'].apply(lambda x: x.split('.')[3] if len(x.split('.')) > 3 else '')


sorted_df = df.groupby('4th_value').apply(shuffle_group).reset_index(drop=True)


sorted_df = sorted_df.drop(columns=['4th_value'])


sorted_df = pd.concat([sorted_df, sum_row])


wb = load_workbook(input_file)
ws = wb[sheet_name]


row_styles = {}
for row in ws.iter_rows(min_row=6, max_row=ws.max_row, max_col=ws.max_column):
    row_styles[row[0].row] = [(cell.value, cell._style) for cell in row]

ws.delete_rows(6, ws.max_row - 5)



for i, row in enumerate(dataframe_to_rows(sorted_df, index=False, header=False), start=6):
    for j, value in enumerate(row, start=1):
        cell = ws.cell(row=i, column=j, value=value)
        if i in row_styles:
           
            cell._style = row_styles[i][j-1][1]

wb.save(output_file)

print(f"Sorting and shuffling completed. The sorted data is saved to {output_file}.")